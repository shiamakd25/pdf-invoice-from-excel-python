import pandas as pd
import glob
import openpyxl
from fpdf import FPDF
from pathlib import Path

files = glob.glob("invoices/*.xlsx")

for file in files:
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()

    file_name = Path(file).stem
    invoice_num = file_name.split("-")[0]
    date=file_name.split("-")[1]

    pdf.set_font(family="Times", size=16, style="B")
    pdf.cell(w=50, h=8, txt=f"Invoice Number {invoice_num}", ln=1)
    pdf.cell(w=50, h=8, txt=f"Date: {date}", ln=1)
    pdf.ln(10)

    df = pd.read_excel(file, sheet_name="Sheet 1")

    columns = list(df.columns)
    columns = [item.replace("_", " ").title() for item in columns]

    pdf.set_font(family="Times", size=10, style="B")
    pdf.cell(w=70, h=8, txt=columns[0], border=1)
    pdf.cell(w=30, h=8, txt=columns[1], border=1)
    pdf.cell(w=40, h=8, txt=columns[2], border=1)
    pdf.cell(w=30, h=8, txt=columns[3], border=1)
    pdf.cell(w=20, h=8, txt=columns[4], border=1, ln=1)

    for index, row in df.iterrows():
        pdf.set_font(family="Times", size=10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(w=70, h=8, txt=str(row["product_name"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["product_id"]), border=1)
        pdf.cell(w=40, h=8, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=30, h=8, txt=f"${str(row['price_per_unit'])}", border=1)
        pdf.cell(w=20, h=8, txt=f"${str(row['total_price'])}", border=1, ln=1)

    total_sum = df["total_price"].sum()

    pdf.set_font(family="Times", size=10, style="B")
    pdf.set_text_color(0, 0, 0)
    pdf.cell(w=70, h=8, txt="")
    pdf.cell(w=30, h=8, txt="")
    pdf.cell(w=40, h=8, txt="")
    pdf.cell(w=30, h=8, txt="")
    pdf.cell(w=20, h=8, txt=f"${str(total_sum)}", border=1, ln=1)

    pdf.ln(10)

    pdf.set_font(family="Times", size=12)
    pdf.cell(w=30, h=8, txt=f"The total price is ${total_sum}", ln=1)

    pdf.set_font(family="Times", size=12, style="B")
    pdf.cell(w=30, h=8, txt="Shiamak Das")

    pdf.output(f"pdfs/{file_name}.pdf")
